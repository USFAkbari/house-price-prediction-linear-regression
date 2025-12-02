#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ایجاد فایل اکسل با فرمول‌های محاسبه قیمت
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# خواندن پارامترهای مدل
with open('model_params.json', 'r', encoding='utf-8') as f:
    params = json.load(f)

feature_names = params['feature_names']
coefficients = params['coefficients']
intercept = params['intercept']
means = params['means']
stds = params['stds']

# ایجاد workbook
wb = Workbook()
ws = wb.active
ws.title = "محاسبه قیمت"

# استایل‌ها
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
input_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
result_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
result_font = Font(bold=True, size=14)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ردیف 1: هدرها
ws['A1'] = "ویژگی"
ws['B1'] = "مقدار ورودی"
ws['C1'] = "میانگین"
ws['D1'] = "انحراف معیار"
ws['E1'] = "مقدار نرمال‌سازی شده"
ws['F1'] = "ضریب"
ws['G1'] = "ضریب × نرمال‌سازی شده"

for col in range(1, 8):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# ردیف 2: پارامترهای مدل (برای استفاده در فرمول‌ها)
# این ردیف مخفی خواهد بود
for i, (name, mean, std, coef) in enumerate(zip(feature_names, means, stds, coefficients), start=2):
    ws.cell(row=i, column=1, value=name).border = border
    ws.cell(row=i, column=3, value=mean).border = border
    ws.cell(row=i, column=4, value=std).border = border
    ws.cell(row=i, column=6, value=coef).border = border
    
    # سلول ورودی (B)
    input_cell = ws.cell(row=i, column=2)
    input_cell.fill = input_fill
    input_cell.border = border
    input_cell.alignment = Alignment(horizontal='center')
    
    # فرمول نرمال‌سازی (E): (B - C) / D
    norm_cell = ws.cell(row=i, column=5)
    norm_cell.value = f"=IF(ISBLANK(B{i}),0,(B{i}-C{i})/D{i})"
    norm_cell.fill = formula_fill
    norm_cell.border = border
    norm_cell.alignment = Alignment(horizontal='center')
    
    # فرمول ضریب × نرمال‌سازی شده (G): F * E
    contrib_cell = ws.cell(row=i, column=7)
    contrib_cell.value = f"=F{i}*E{i}"
    contrib_cell.fill = formula_fill
    contrib_cell.border = border
    contrib_cell.alignment = Alignment(horizontal='center')

# ردیف بعدی: Intercept
intercept_row = len(feature_names) + 2
ws.cell(row=intercept_row, column=1, value="Intercept (قیمت پایه)").font = Font(bold=True)
ws.cell(row=intercept_row, column=2, value=intercept).fill = formula_fill
ws.cell(row=intercept_row, column=2).border = border

# ردیف بعدی: مجموع
sum_row = intercept_row + 1
ws.cell(row=sum_row, column=1, value="مجموع تأثیرات").font = Font(bold=True)
sum_cell = ws.cell(row=sum_row, column=2)
sum_cell.value = f"=SUM(G2:G{len(feature_names)+1})"
sum_cell.fill = formula_fill
sum_cell.border = border
sum_cell.alignment = Alignment(horizontal='center')

# ردیف نهایی: قیمت نهایی
price_row = sum_row + 1
ws.cell(row=price_row, column=1, value="قیمت پیش‌بینی شده").font = Font(bold=True, size=14)
price_cell = ws.cell(row=price_row, column=2)
price_cell.value = f"=B{intercept_row}+B{sum_row}"
price_cell.fill = result_fill
price_cell.font = result_font
price_cell.border = border
price_cell.alignment = Alignment(horizontal='center', vertical='center')

# تنظیم عرض ستون‌ها
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 30

# مخفی کردن ردیف پارامترها (اختیاری - می‌توانیم آن‌ها را نمایش دهیم)
# ws.row_dimensions[2].hidden = True  # اگر بخواهیم مخفی کنیم

# ایجاد شیت دوم: پارامترهای مدل
ws2 = wb.create_sheet("پارامترهای مدل")

# هدرها
headers = ["ویژگی", "میانگین", "انحراف معیار", "ضریب رگرسیون"]
for col, header in enumerate(headers, start=1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# داده‌ها
for i, (name, mean, std, coef) in enumerate(zip(feature_names, means, stds, coefficients), start=2):
    ws2.cell(row=i, column=1, value=name).border = border
    ws2.cell(row=i, column=2, value=mean).border = border
    ws2.cell(row=i, column=3, value=std).border = border
    ws2.cell(row=i, column=4, value=coef).border = border

# Intercept
intercept_row2 = len(feature_names) + 2
ws2.cell(row=intercept_row2, column=1, value="Intercept").font = Font(bold=True)
ws2.cell(row=intercept_row2, column=2, value=intercept).font = Font(bold=True)
ws2.cell(row=intercept_row2, column=2).border = border

# تنظیم عرض ستون‌ها
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18

# اضافه کردن توضیحات
desc_row = intercept_row2 + 2
ws2.cell(row=desc_row, column=1, value="توضیحات:").font = Font(bold=True)
ws2.cell(row=desc_row+1, column=1, value="فرمول محاسبه قیمت:")
ws2.cell(row=desc_row+2, column=1, value="Price = Intercept + Σ(Coefficient_i × Normalized_Feature_i)")
ws2.cell(row=desc_row+3, column=1, value="")
ws2.cell(row=desc_row+4, column=1, value="فرمول نرمال‌سازی:")
ws2.cell(row=desc_row+5, column=1, value="Normalized_Feature = (Feature_Value - Mean) / Std")

# ذخیره فایل
wb.save('price_calculator.xlsx')
print("فایل 'price_calculator.xlsx' با موفقیت ایجاد شد!")
print(f"\nبرای استفاده:")
print(f"1. در شیت 'محاسبه قیمت'، مقادیر ویژگی‌ها را در ستون B وارد کنید")
print(f"2. قیمت به صورت خودکار در ردیف {price_row} محاسبه می‌شود")

